import os, sys
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal
from helpers import convert_rate, convert_sheet
import re

income_types = ["Interest received", "Interest received from loan repurchase", "Late fees received", "Delayed interest income on transit rebuy", "Interest received from pending payments"]
cost_types = []
witholding_tax = Decimal("0.05")
polish_tax = Decimal("0.19")

def parse_isin_loan(text):
    match = re.search(r"ISIN: (\w+) \(Loan (\d+-\d+)\)", text)
    if match:
        return match.group(1) + '-' + match.group(2)
    return None

def process_transactions(path):
    workbook = load_workbook(filename=path, read_only=False)
    transactions = dict()
    withloding_taxes = Decimal('0')

    for s in workbook.sheetnames:

        sheet = convert_sheet(workbook[s])
        for row in sheet:
            if row['Date'] is None:
                continue
            trans_type = row['Payment Type']
            date = datetime.strptime(row['Date'], '%Y-%m-%d %H:%M:%S')
            amount = Decimal(str(row["Turnover"]))
            isin_loan = parse_isin_loan(row['Details'])
            if isin_loan not in transactions:
                transactions[isin_loan] = []

            trans_group = transactions[isin_loan]
            exiting_trans = next((x for x in trans_group if x['date'] == date), None)
            if trans_type in income_types:
                if exiting_trans is not None:
                    exiting_trans['amount'] += amount
                else:
                    trans_group.append({'isin_loan': isin_loan, 'type': 'profit', 'subtype': trans_type, 'date': date, 'amount': amount, 'currency': row['Currency']})
            elif trans_type == 'Tax withholding':
                withloding_taxes += convert_rate(date, amount, row['Currency'])
            elif trans_type in cost_types:
                trans_group.append({'isin_loan': isin_loan, 'type': 'fee', 'date': date, 'amount': amount, 'currency':  row['Currency']})
            else:
                print(f"Unknown transaction type {trans_type}")
                exit(1)

    return [item for row in transactions.values() for item in row], withloding_taxes


def calculate_tax(path):
    przychod = Decimal("0")
    cost = Decimal("0")
    total_tax = Decimal('0')

    transactions, withloding_taxes = process_transactions(path)
    for trans in transactions:
        if trans['type'] == 'profit':
            przychod += convert_rate(trans['date'], trans['amount'], trans['currency'])
        elif trans['type'] == 'fee':
            if trans['wth'] is not None:
                raise Exception('{trans} feew with witholding tax??')
            cost += convert_rate(trans['date'], trans['amount'], trans['currency'])
        else:
            raise Exception('wtf')

    total_tax = round(total_tax, 2)
    przychod = round(przychod, 2)
    cost = round(cost, 2)

    dochod = przychod - cost
    total_tax = round(polish_tax * przychod, 2)
    tax_paid_abroad = -round(withloding_taxes, 2)
    tax_to_pay = total_tax - tax_paid_abroad

    return przychod, dochod, cost, total_tax, tax_paid_abroad, tax_to_pay

przychod, dochod, cost, total_tax, tax_paid_abroad, tax_to_pay = calculate_tax('mintos.xlsx')

print("Mintos rozliczamy w PIT-38!")
print(f"Przychód w pln: {przychod} zł")
print(f"Dochód w pln: {dochod} zł")
print(f"Koszt w pln: {cost} zł")
print(f"Podatek łącznie: {total_tax} zł")
print(f"Podatek zaplacony: {tax_paid_abroad} zł")
print(f"Podatek do zapłacenia: {tax_to_pay} zł  TO WPISUJEMY DO PITA") 