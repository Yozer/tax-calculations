import os, sys
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal
from helpers import convert_rate, convert_sheet, warsaw_timezone

tax_rate = Decimal("0.19")
income_types = []
cost_types = []

trans_types_to_ignore = ['Interest', 'Exchange To Withdraw', 'Fixed Term Interest', 'Unlocking Term Deposit']

def calculate_tax(path):
    workbook = load_workbook(filename=path, read_only=False)
    income = Decimal("0")
    cost = Decimal("0")

    sheet = convert_sheet(workbook[workbook.sheetnames[0]])
    for row in sheet:
        if row['Date / Time (UTC)'] is None:
            continue
        trans_type = row['Type']
        if trans_type in trans_types_to_ignore:
            continue

        input_currency = row['Input Currency']
        output_currency = row['Output Currency']
        amount = Decimal(str(row["Input Amount"]))
        date = row['Date / Time (UTC)'] if isinstance(row['Date / Time (UTC)'], datetime) else datetime.strptime(row['Date / Time (UTC)'], '%Y-%m-%d %H:%M:%S').astimezone(warsaw_timezone)

        if trans_type == 'Withdraw Exchanged':
            if output_currency == 'EUR' or 'USD':
                income += convert_rate(date, amount, output_currency)
            else:
                raise Exception("invalid output currency")
        elif trans_type == 'Exchange Deposited On':
            if input_currency == 'EUR' or 'USD':
                cost += convert_rate(date, amount, input_currency)
            else:
                raise Exception("invalid output currency")

    return round(income, 2), round(cost, 2)

income, cost = calculate_tax('nexo.xlsx')
dochod = max(income - cost, 0)
tax = dochod * tax_rate

print("W KRYPTO TO WPISUJEMY!")
print("SPRAWDZIC PRZY SPRZEDAWANIU EURX I INNYCH CRYPTO w 2025")
print(f"Przychód w pln: {income} zł")
print(f"Dochód w pln: {dochod} zł")
print(f"Koszt w pln: {cost} zł")
print(f"Podatek: ~{tax} zł")