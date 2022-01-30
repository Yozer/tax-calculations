import os, sys, re
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from openpyxl import load_workbook
import datetime
from helpers import convert_rate, convert_sheet, warsaw_timezone, fiat_currencies
from decimal import Decimal

operations_to_skip = []
operations_to_process = ["MARKET_SELL", "LIMIT_SELL", "CEILING_MARKET_BUY", "LIMIT_BUY"]
excel_date_format = '%m/%d/%Y %I:%M:%S %p'

def calculate_tax():
    file_name = 'bittrex.xlsx'

    if not os.path.exists(file_name):
        print(f'WARNING: Bittrex {file_name} doesnt exist. Skipping')
        return

    workbook = load_workbook(filename=file_name)
    transactions = convert_sheet(workbook[workbook.sheetnames[0]])

    przychod_total = Decimal(0)
    koszt_total = Decimal(0)
    fiat_staking_total = Decimal(0)

    for row in transactions:
        if row['Uuid'] is None:
            continue

        type = row['OrderType']
        if type in operations_to_skip:
            continue
        if type not in operations_to_process:
            raise Exception(f'Bittrex. Unknown transaction type {type}')

        asOfDate = datetime.datetime.strptime(row['Closed'], excel_date_format)
        trade_id = row["Uuid"]
        price = Decimal(str(row["Price"]))
        fee = Decimal(str(row["Commission"]))
        ticker = row["Exchange"].split('-')
        if len(ticker) != 2:
            raise Exception(f'Bittrex invalid ticker for {trade_id}')

        pln = convert_rate(asOfDate, price, currency='EUR')
        if ticker[0] == 'EUR':
            koszt_total += convert_rate(asOfDate, fee, currency='EUR')
            if 'SELL' in type:
                przychod_total += pln
            else:
                koszt_total += pln
        elif ticker[1] == 'EUR':
            raise Exception('Bittrex: Untested')
            # if 'SELL' in type:
            #     koszt_total += pln
            # else:
            #     przychod_total += pln

    return ("Bittrex", round(przychod_total, 2), round(koszt_total, 2), round(fiat_staking_total, 2))