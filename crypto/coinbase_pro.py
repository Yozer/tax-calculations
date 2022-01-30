import os, sys, re
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from openpyxl import load_workbook
from helpers import convert_rate, convert_sheet, warsaw_timezone, fiat_currencies
from decimal import Decimal

operations_to_skip = ["deposit", "withdrawal"]
operations_to_process = ["match", "fee"]

def calculate_tax():
    file_name = 'coinbase_pro.xlsx'

    if not os.path.exists(file_name):
        print(f'WARNING: Coinbase pro {file_name} doesnt exist. Skipping')
        return

    workbook = load_workbook(filename=file_name)
    transactions = convert_sheet(workbook[workbook.sheetnames[0]])

    przychod_total = Decimal(0)
    koszt_total = Decimal(0)
    fiat_staking_total = Decimal(0)

    for row in transactions:
        if row['portfolio'] is None:
            continue

        type = row['type']
        if type in operations_to_skip:
            continue
        if type not in operations_to_process:
            raise Exception(f'Coinbase pro. Unknown transaction type {type}')

        asOfDate = row["time"]
        trade_id = row["trade id"]
        amount = Decimal(str(row["amount"]))
        coin = row["amount/balance unit"]

        if type == 'fee' and coin in fiat_currencies:
            if amount >= 0:
                raise Exception(f"Positive fee for trade_id: {trade_id}")
            koszt_total -= convert_rate(asOfDate, amount, currency=coin)

        if type != 'match' or coin not in fiat_currencies:
            continue
        pln = convert_rate(asOfDate, amount, currency=coin)
        if pln > 0:
            przychod_total += pln
        else:
            koszt_total -= pln

    return ("Coinbase PRO", round(przychod_total, 2), round(koszt_total, 2), round(fiat_staking_total, 2))