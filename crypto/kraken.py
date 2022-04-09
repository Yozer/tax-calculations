import os, sys, re
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from openpyxl import load_workbook
from helpers import convert_rate, convert_sheet
from decimal import Decimal

operations_to_skip = ["deposit", "withdrawal", "transfer"]
operations_to_process = ["staking", "trade", "spend", "receive"]

def calculate_tax():
    file_name = 'kraken.xlsx'

    if not os.path.exists(file_name):
        print(f'WARNING: Kraken {file_name} doesnt exist. Skipping')
        return(None, None, None, None)

    workbook = load_workbook(filename=file_name)
    transactions = convert_sheet(workbook[workbook.sheetnames[0]])

    przychod_total = Decimal(0)
    koszt_total = Decimal(0)
    fiat_staking_total = Decimal(0)

    for row in transactions:
        if row['time'] is None:
            continue

        type = row["type"]
        if type in operations_to_skip:
            continue
        if type not in operations_to_process:
            raise Exception(f'Unkown operation for Kraken: {type}')

        txid = row["txid"]
        amount = Decimal(str(row['amount']))
        fee = Decimal(str(row['fee']))
        asOfDate = row["time"]
        if row['asset'] not in ['ZEUR', 'EUR.M']:
            continue

        pln_amount = convert_rate(asOfDate, amount, currency='EUR')
        fee_pln = convert_rate(asOfDate, fee, currency='EUR')

        if type == 'staking':
            fiat_staking_total += pln_amount
        elif type == 'spend':
            if pln_amount >= 0:
                raise Exception(f"Kraken: positive spend transaction for txin: {txid}")
            if not fee.is_zero():
                raise Exception(f"Kraken: positive fee for spend transaction for txin: {txid}")
            koszt_total -= pln_amount
        elif type == "receive":
            if pln_amount <= 0:
                raise Exception(f"Kraken: negative receive transaction for txin: {txid}")
            if not fee.is_zero():
                raise Exception(f"Kraken: positive fee for receive transaction for txin: {txid}")

            przychod_total += pln_amount

        elif type == "trade":
            if not fee.is_zero():
                koszt_total += fee_pln
            if pln_amount > 0:
                przychod_total += pln_amount
            else:
                koszt_total -= pln_amount


    return ("Kraken", round(przychod_total, 2), round(koszt_total, 2), round(fiat_staking_total, 2))