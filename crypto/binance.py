import os, sys, re
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from openpyxl import load_workbook
from helpers import convert_rate, convert_sheet, warsaw_timezone, fiat_currencies
from datetime import datetime
from decimal import Decimal

operations_to_skip = ["Deposit", "Withdraw", "Savings purchase", "Savings Principal redemption", "transfer_out", "transfer_in", "Binance Card Spending", "Fiat Deposit"]
operations_to_process = ["Transaction Related", "Savings Interest", "Sell", "Fee", "Distribution", "Transaction Buy", "Transaction Sold"]
ignored_coins = set()

def calculate_tax():
    file_name = 'binance.xlsx'

    if not os.path.exists(file_name):
        print(f'WARNING: Binance {file_name} doesnt exist. Skipping')
        return(None, None, None, None)

    workbook = load_workbook(filename=file_name)
    sheet = convert_sheet(workbook[workbook.sheetnames[0]])

    przychod_total = Decimal(0)
    koszt_total = Decimal(0)
    fiat_staking_total = Decimal(0)

    for row in sheet:
        if row["Account"] == "Card" or row["User_ID"] is None:
            continue
        coin = row["Coin"]
        if coin not in fiat_currencies:
            if coin in ignored_coins:
                continue
            print(f"BINANCE: Ignoring coin: '{coin}' It's fine as long as it's not fiat")
            ignored_coins.add(coin)
            continue

        if row["Account"] not in ["SPOT", "Savings", "CARD"]:
            raise Exception(f"Unknown account type for Binance {row['Account']}")

        operation = row["Operation"]
        if operation in operations_to_skip:
            continue
        if operation not in operations_to_process:
            raise Exception(f'Unkown operation for Binance: {operation}')

        asOfDate = datetime.strptime(row['UTC_Time'], '%Y-%m-%d %H:%M:%S').astimezone(warsaw_timezone)
        change = Decimal(str(row["Change"]))
        pln = round(convert_rate(asOfDate, change, currency=coin), 2)

        if operation in ["Distribution", "Savings Interest"]:
            fiat_staking_total += pln
        elif operation == "Fee":
            if change >= 0:
                raise Exception(f"Found positive fee {change}")
            koszt_total -= pln
        elif change < 0:
            koszt_total -= pln
        else:
            przychod_total += pln

    return ("Binance", przychod_total, koszt_total, fiat_staking_total)