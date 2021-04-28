import re
from openpyxl import load_workbook
from datetime import datetime, timedelta
from itertools import groupby
from decimal import Decimal
from helpers import sum_dict, convert_rate, convert_sheet, group_by_pos_id, add_working_days
from mapping import mapping

# pos_types: crypto, stock, dividend, fee
tax_rate = Decimal("0.19")
crypto = set(["BTC/USD", "ETH/USD", "BCH/USD", "XRP/USD", "DASH/USD", "LTC/USD", "ETC/USD", "ADA/USD", "IOTA/USD", "MIOTA/USD", "XLM/USD", "EOS/USD", "NEO/USD", "TRX/USD", "ZEC/USD", "BNB/USD", "XTZ/USD"])
ignored_transactions = ['Deposit', 
                        'Start Copy', 
                        'Account balance to mirror', 
                        'Mirror balance to account',
                        'Stop Copy',
                        'Edit Stop Loss']
traded_cryptos = set()
unknown_stocks = set()
CryptoCountry = 'CryptoCountry'
UnknownCountry = 'Unknown'

def get_position_type(pos_id, transactions, closed_positions):
    if pos_id not in transactions:
        raise Exception(f'Logic error. Unable to find position {pos_id} in transactions sheet')

    first_transaction = transactions[pos_id][0] # we might have different types
    stock_name = first_transaction['Details']
    is_cfd = closed_positions[pos_id][0]["Is Real"] == "CFD"
    pos_type = "crypto" if stock_name in crypto and not is_cfd else "stock"
    if pos_type == "crypto":
        traded_cryptos.add(stock_name)

    return pos_type

def get_mapped_country(stock):
    for k, v in mapping.items():
        if stock in v:
            return k

    if stock.endswith("/USD") and "." not in stock:
        return "USA"
    unknown_stocks.add(stock)
    return UnknownCountry

def get_ticker_country(pos_id, transactions, closed_positions, dividend=False):
    if pos_id not in transactions:
        raise Exception(f'Logic error. Unable to find position {pos_id} in transactions sheet')
    
    stock = transactions[pos_id][0]["Details"]
    if dividend:
        return "DividendCountry"

    closed_position = closed_positions[pos_id][0]
    is_cfd = closed_position["Is Real"] == "CFD"
    if is_cfd:
        return "Malta"

    pos_type = get_position_type(pos_id, transactions, closed_positions)
    if pos_type == "crypto":
        return CryptoCountry
    
    return get_mapped_country(stock)

def process_rollover_fee(transaction, transactions, closed_positions):
    amount = Decimal(str(transaction["Amount"]))
    pos_id = transaction["Position ID"]
    date = datetime.strptime(transaction['Date'], '%Y-%m-%d %H:%M:%S')
    equity_change = Decimal(str(transaction['Realized Equity Change']))

    if amount != equity_change:
        raise Exception(f'Weird row on position id {transaction["Position ID"]}. Closing')

    if transaction['Details'] in ['Weekend fee', 'Over night fee']:
        pos_type = 'fee'
        country = get_ticker_country(pos_id, transactions, closed_positions)
        if country == CryptoCountry:
            # 840652308
            raise Exception(f"Found a rollover fee for crypto position {pos_id}. Should be marked as cfd?")
    elif transaction['Details'] == 'Payment caused by dividend':
        pos_type = 'dividend'
        country = get_ticker_country(pos_id, transactions, closed_positions, True)
        if country == CryptoCountry:
            raise Exception(f"Found a dvidend for crypto position {pos_id}. Should be marked as cfd?")
    else:
        raise Exception(f"Unkown fee {row['Details']} for position {row['Position ID']}")

    return ({'id': pos_id, 'date': date, 'amount': amount, 'country': country, "type": pos_type})

def read(path):
    workbook = load_workbook(filename=path, read_only=False)
    # wb2 = load_workbook('statement_2020.xlsx')
    transactions = convert_sheet(workbook['Transactions Report'])
    grouped_transactions = group_by_pos_id(transactions)
    closed_positions = convert_sheet(workbook['Closed Positions'])
    grouped_closed_positions = group_by_pos_id(closed_positions)
    entries = []

    for row in closed_positions:
        pos_id = row['Position ID']
        if pos_id is None:
            continue
        amount = Decimal(str(row['Amount']).replace(',', '.'))
        profit = Decimal(str(row['Profit']).replace(',', '.'))
        units = Decimal(row['Units'].replace(',', '.'))
        leverage = Decimal("1") if row['Leverage'] == None else Decimal(str(row['Leverage']).replace(',', '.'))
        pos_type = get_position_type(pos_id, grouped_transactions, grouped_closed_positions)
        is_cfd = row["Is Real"] == "CFD"
        # d = 2 if pos_type == "stock" and not is_cfd else 0
        d = 0
        trans = {
            'open_date': add_working_days(datetime.strptime(row['Open Date'], '%d-%m-%Y %H:%M'), d),
            'close_date': add_working_days(datetime.strptime(row['Close Date'], '%d-%m-%Y %H:%M'), d),
            'open_amount': amount * leverage,
            'close_amount': amount * leverage,
            'is_cfd': is_cfd,
            'id': pos_id,
            'type': pos_type,
            'status': 'closed',
            'country': get_ticker_country(pos_id, grouped_transactions, grouped_closed_positions)
        }

        if row["Action"].startswith("Sell"):
            trans["close_amount"] -= profit
            trans["close_amount"], trans["open_amount"] = trans["open_amount"], trans["close_amount"]
            trans["open_date"], trans["close_date"] = trans["close_date"], trans["open_date"]
        else:
            trans["close_amount"] += profit

        # if trans['close_date'].year != 2020:
        #     continue
        entries.append(trans)

    for row in transactions:
        pos_id = row['Position ID']
        if pos_id is None:
            continue
        trans_type = row['Type']
        if trans_type == 'Rollover Fee':
            entries.append(process_rollover_fee(row, grouped_transactions, grouped_closed_positions))
        elif trans_type == "Open Position":
            if pos_id not in grouped_closed_positions:
                if row["Details"] in crypto:
                    print(f"Found krypto that was bought but not sold. Unable to determine CFD. Assuming not. Veryify {pos_id}")
                else:
                    continue
                
                trans = {
                    'open_date': add_working_days(datetime.strptime(row['Date'], '%Y-%m-%d %H:%M:%S'), d),
                    'close_date': add_working_days(datetime.strptime(row['Date'], '%Y-%m-%d %H:%M:%S'), d),
                    'open_amount': Decimal(str(row["Amount"])),
                    'close_amount': Decimal("0"),
                    'is_cfd': False,
                    'id': pos_id,
                    'type': 'crypto',
                    'status': 'open',
                    'country': CryptoCountry
                }
                entries.append(trans)
        elif trans_type == "Profit/Loss of Trade":
            if pos_id not in grouped_closed_positions:
                raise Exception(f"Closed but not in closed? wtf {pos_id}")
        elif trans_type not in ignored_transactions:
            raise Exception(f'Unknown transaction type {trans_type} for position {pos_id}')

    return entries

def process_positions(positions, typ):
    positions = [x for x in positions if x["type"] == typ or (typ == "stock" and x['type'] == "fee")]
    income_usd = {}
    przychod = {}
    koszty = {}
    dochod = {}

    for pos in positions:
        country = pos["country"]
        if country not in income_usd:
            income_usd[country] = Decimal("0")
            przychod[country] = Decimal("0")
            koszty[country] = Decimal("0")
            dochod[country] = Decimal("0")

        if pos["type"] == "fee":
            rate_pln = convert_rate(pos["date"], pos["amount"])
            koszty[country] -= rate_pln
            dochod[country] += rate_pln
        else:
            if pos["status"] == "closed":
                income_usd[country] += pos["close_amount"] - pos["open_amount"]
            open_rate_pln = convert_rate(pos["open_date"], pos["open_amount"])
            close_rate_pln = convert_rate(pos["close_date"], pos["close_amount"])
            przychod[country] += close_rate_pln
            koszty[country] += open_rate_pln
            dochod[country] += close_rate_pln - open_rate_pln

    return (income_usd, przychod, koszty, dochod)

fname = 'statement_2020.xlsx'
# fname = 'statement_2020 - Copy.xlsx'
entries = read(fname)
income_stock_usd, przychod_stock, koszty_stock, dochod_stock = process_positions(entries, 'stock')
print()
print(f"Dochód $ za stocks: ${sum_dict(income_stock_usd)}")
print(f"Przychód w pln za stocks: {sum_dict(przychod_stock)} zł")
print(f"Koszt w pln za stocks: {sum_dict(koszty_stock)} zł")
print(f"Dochód w pln za stocks: {sum_dict(dochod_stock)} zł")
print(f'Dochód per kraj: {dict([(x, str(y)) for x, y in dochod_stock.items()])}')

if len(unknown_stocks) > 0:
    print(f'Unkown stocks: {unknown_stocks}')

income_crypto_usd, przychod_crypto, koszty_crypto, dochod_crypto = process_positions(entries, 'crypto')
print()
print(f"Cryptos: {list(traded_cryptos)}")
print(f"Dochód $ za crypto: ${sum_dict(income_crypto_usd)}")
print(f"Przychód w pln za crypto: {sum_dict(przychod_crypto)} zł")
print(f"Koszt w pln za crypto: {sum_dict(koszty_crypto)} zł")
print(f"Dochód w pln za crypto: {sum_dict(dochod_crypto)} zł")