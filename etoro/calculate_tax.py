import os, sys
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from openpyxl import load_workbook
from datetime import datetime, timedelta
from decimal import Decimal
from mapping import get_country_code, CryptoCountry, CfdCountry
from helpers import sum_dict, convert_rate, convert_sheet

# pos_types: crypto, stock, dividend, fee
CryptoType = 'crypto'
StockType = 'stock'
DividendType = 'dividend'
FeeType = 'fee'
InterestType = 'interest'
AdjustmentType = 'adjustment' # etoro made a mistake and either took from us or gave us
IndexAdjustmentType = 'index-adjustment' # no idea yet what it is
RefundType = 'refund'

tax_rate = Decimal("0.19")
use_t_plus_2 = False
year = 2024
ignored_transactions = ['Deposit',
                        'Start Copy',
                        'Account balance to mirror',
                        'Mirror balance to account',
                        'Withdraw Request',
                        'Stop Copy',
                        'Edit Stop Loss',
                        'corp action: Split',
                        'AirDrop',
                        'Staking']
excel_date_format = '%d/%m/%Y %H:%M:%S'

def t2_date(date: datetime):
    if not use_t_plus_2:
        return date

    num_days = 2
    while num_days > 0:
        date += timedelta(days=1)
        if date.weekday() < 5: # Monday to Friday are considered weekdays
            num_days -= 1

    if date.day == 1 and date.month == 1:
        date += timedelta(days=1)
    return date

def group_by_pos_id(transactions):
    res = {}
    for x in transactions:
        pos_id = x["Position ID"]
        if pos_id not in res:
            res[pos_id] = []
        res[pos_id].append(x)

    return res

def get_ticker_country(position, transactions, closed_positions, dividends):
    pos_id = position['id']
    if pos_id not in transactions:
        raise Exception(f'Logic error. Unable to find position {pos_id} in transactions sheet')

    if position['type'] == CryptoType:
        return CryptoCountry
    elif position['is_cfd']:
        return CfdCountry
    elif position['type'] not in [StockType, FeeType]:
        raise Exception(f'Unexpected position type {position["type"]} for {pos_id}')

    first_transaction = next((t for t in transactions[pos_id] if t['Type'] in ['Position closed', 'Open Position', 'Dividend']))
    closed_position = closed_positions[pos_id][0] if pos_id in closed_positions else None
    stock_name = None if closed_position is None else closed_position["Action"]
    stock_symbol = first_transaction['Details']

    if closed_position is None:
        dividend = dividends[pos_id][0] if pos_id in dividends else None
        stock_name = None if dividend is None else dividend["Instrument Name"]

    return get_country_code(stock_name, stock_symbol)

def parse_decimal(r):
    return Decimal(str(r))
    
def parse_date(r):
    return datetime.strptime(r, excel_date_format)

def is_asset_cfd(r):
    if r['Asset type'] in ['', None]:
        raise Exception(f'Empty asset type for {r["pos_id"]}')
    return r['Asset type'] == 'CFD'

def process_interest_payment(transaction):
    pos_id = transaction["Position ID"]
    amount = parse_decimal(transaction["Amount"])
    date = parse_date(transaction['Date'])
    trans = {'id': pos_id, 'date': date, 'amount': amount, "type": InterestType, 'equity_change': amount, "is_cfd": True}
    return trans

def process_adjustment(transaction):
    pos_id = transaction["Position ID"]
    amount = parse_decimal(transaction["Amount"])
    date = parse_date(transaction['Date'])
    if transaction['Type'] == 'Index price adjustment':
        type = IndexAdjustmentType
    else:
        type = AdjustmentType
    trans = {'id': pos_id, 'date': date, 'amount': amount, "type": type, 'equity_change': amount, "is_cfd": True}
    return trans

def process_rollover_fee(transaction):
    amount = parse_decimal(transaction["Amount"])
    pos_id = transaction["Position ID"]
    date = parse_date(transaction['Date'])

    transaction_details = transaction['Details'].lower()
    transaction_type = transaction['Type'].lower()
    if (transaction_type == 'overnight fee' and transaction_details in ['weekend fee', 'daily']) or transaction_type == 'sdrt':
        pos_type = FeeType
    elif transaction_type == 'dividend':
        pos_type = DividendType
    elif (transaction_type == 'overnight refund' and transaction_details == 'daily') or (transaction_type == 'weekend refund' and transaction_details == 'weekend fee'):
        if amount < 0:
            raise Exception(f"Negative refund {transaction_details} for position {transaction['Position ID']}")
        pos_type = RefundType
    else:
        raise Exception(f"Unkown fee {transaction_details} for position {transaction['Position ID']}")

    return {'id': pos_id, 'date': date, 'amount': amount, "type": pos_type, "is_cfd": is_asset_cfd(transaction)}

def read_dividend_taxes(path):
    workbook = load_workbook(filename=path)
    sheet = convert_sheet(workbook['Dividends'])
    dividend_taxes = {}
    for x in sheet:
        pos_id = str(x["Position ID"])
        if x["Position ID"] is None:
            continue
        if pos_id not in dividend_taxes:
            dividend_taxes[pos_id] = []

        x["Withholding Tax Rate (%)"] = parse_decimal(x["Withholding Tax Rate (%)"].replace('%', '')) / Decimal("100")
        x["Net Dividend Received (USD)"] = parse_decimal(x["Net Dividend Received (USD)"])
        x["Withholding Tax Amount (USD)"] = parse_decimal(x["Withholding Tax Amount (USD)"])
        x["Date of Payment"] = datetime.strptime(str(x["Date of Payment"]), '%d/%m/%Y')

        if x["Date of Payment"].year != year:
            raise Exception('Invalid year found in excel')
        dividend_taxes[pos_id] += [x]

    return (dividend_taxes, sheet)

def get_asset_type(asset):
    asset = asset['Asset type']
    if asset in ['Stocks', 'ETF', 'CFD']:
        return StockType
    elif asset == 'Crypto':
        return CryptoType
    else:
        raise Exception(f'Failed to parse {asset}')

def read(path):
    workbook = load_workbook(filename=path)
    transactions = convert_sheet(workbook['Account Activity'])
    grouped_transactions = group_by_pos_id(transactions)
    closed_positions = convert_sheet(workbook['Closed Positions'])
    grouped_closed_positions = group_by_pos_id(closed_positions)
    entries = []

    # sanity checks... because etoro does more bugs than I
    has_fatal_errors = False
    for row  in closed_positions:
        pos_id = row['Position ID']
        if pos_id is None:
            continue

        close_date = parse_date(row['Close Date'])
        open_date = parse_date(row['Open Date'])

        is_cfd = row["Type"] == "CFD"
        is_sell = row["Action"].startswith("Sell")        
        if is_sell and not is_cfd:
            raise Exception(f"Not CFD that is sell? Pos id {pos_id}")

        if open_date.year == close_date.year:
            open_activity = next((x for x in grouped_transactions[pos_id] if x['Type'] == 'Position closed'), None)
            if open_activity is None:
                print(f'FATAL ERROR: Missing closed position, report to ETORO! Position id: {pos_id}')
                has_fatal_errors = True

    if has_fatal_errors:
        raise Exception('Aborting due to fatal errors')

    for row in transactions:
        pos_id = row['Position ID']
        if pos_id is None:
            continue

        date = parse_date(row['Date'])
        amount = parse_decimal(row["Amount"])
        trans_type = row['Type']
        asset_type = row['Asset type']
        trans = { 'id': pos_id }
        if date.year != year:
            raise Exception('Invalid year found in excel')

        if trans_type in ['Overnight fee', 'Overnight refund', 'Weekend refund', 'Dividend', 'SDRT']:
            entries.append(process_rollover_fee(row))
        elif trans_type == 'Interest Payment':
            entries.append(process_interest_payment(row))
        elif trans_type in ['Adjustment', 'Index price adjustment']:
            entries.append(process_adjustment(row))
        elif trans_type in ['Withdraw Fee', 'Withdrawal Conversion Fee', 'Deposit Conversion Fee']:
            if amount != 0:
                raise Exception(f'Unsupported withdraw fee/withdrawal conversion fee/depsit conversion fee {amount}')
        elif trans_type == "Open Position":
            # skip as it's taxable only for crypto
            if get_asset_type(row) != CryptoType:
                continue
            if amount <= 0:
                raise Exception(f'Negative crypto buy? {amount}')
            trans['amount'] = -amount
            trans['date'] = date
            trans['type'] = CryptoType
            trans['is_cfd'] = False
            trans['equity_change'] = profit = parse_decimal(row['Realized Equity Change'])
            entries.append(trans)
        elif trans_type == "Position closed":
            profit = parse_decimal(row['Realized Equity Change'])
            parsed_asset_type = get_asset_type(row)
            closed_position = grouped_closed_positions[pos_id][0]
            is_cfd = asset_type == 'CFD'

            if parsed_asset_type != StockType or is_cfd:
                open_date = parse_date(closed_position['Open Date'])
                close_date = parse_date(closed_position['Close Date'])
            else:
                open_date = t2_date(parse_date(closed_position['Open Date']))
                close_date = t2_date(parse_date(closed_position['Close Date']))
            open_amount = parse_decimal(closed_position['Amount'])

            trans['type'] = parsed_asset_type
            trans['equity_change'] = profit

            if len(grouped_closed_positions[pos_id]) > 1:
                raise Exception(f'More than one closed position for {pos_id}')
            if open_amount < 0:
                raise Exception(f'Negative amount for position id {pos_id}')
            if amount < 0: # your dumbass lost more than invested, leverage = bad
                print(f"Check the transaction with negative amount for position id {pos_id}")

            if parsed_asset_type == CryptoType:
                trans['date'] = close_date
                trans['amount'] = amount
                trans['is_cfd'] = False
            elif parsed_asset_type == StockType:
                trans['is_cfd'] = is_cfd
                trans['open_amount'] = open_amount
                trans['close_amount'] = amount
                trans['open_date'] = open_date
                trans['close_date'] = close_date
            else:
                raise Exception(r"Unexpected asset type '{parsed_asset_type}' for {pos_id}")

            entries.append(trans)

        elif trans_type not in ignored_transactions:
            raise Exception(f'Unknown transaction type "{trans_type}" for position {pos_id}')

    return (entries, grouped_transactions, grouped_closed_positions)

def read_summary(path):
    workbook = load_workbook(filename=path)
    summary = convert_sheet(workbook['Financial Summary'])
    stock_sum = Decimal('0')
    crypto_sum = Decimal('0')
    dividends_sum = Decimal('0')
    fees_sum = Decimal('0')
    interest_sum = Decimal('0')
    refunds_sum = Decimal('0')
    index_adjustments_sum = Decimal('0')

    for row in summary:
        amount = parse_decimal(row['Amount\n in (USD)'])

        if row['Name'] in ['CFDs (Profit or Loss)', 'Stocks (Profit or Loss)', 'ETFs (Profit or Loss)']:
            stock_sum += amount
        elif row['Name'] == 'Total Interest payments by eToro EU':
            interest_sum += amount
        elif row['Name'] == 'Crypto (Profit or Loss)':
            crypto_sum += amount
        elif row['Name'] in ['Stock and ETF Dividends (Profit)', 'CFD Dividends (Profit or Loss)']:
            dividends_sum += amount
        elif row['Name'] in ['Fees (overnight, withdrawal, admin)', 'SDRT Charge']:
            fees_sum += amount
        elif row['Name'] == 'Income from Refunds':
            refunds_sum += amount
        elif row['Name'] == 'Index adjustments':
            index_adjustments_sum += amount
        elif row['Name'] in ['Total Return Swaps (Profit or Loss)', 'Income from Airdrops', 'Income from Staking', 'Income from Corporate Actions']:
            if amount != 0:
                raise Exception(f'Unupported: non-zero value for {row["Name"]} in Financial Summary')
        elif row['Name'] in ['Spread fee on CFDs', 'Spread fee on crypto', 'Spread fee on Total Return Swaps (TRS)', 'Spread fee on stocks', 'Spread fee on ETFs']:
            continue
        else:
            raise Exception(f'Unupported: unknown column in {row["Name"]} in Financial Summary')

    return (stock_sum, crypto_sum, dividends_sum, fees_sum, interest_sum, refunds_sum, index_adjustments_sum)

def process_positions(input_positions, typ, unmatched_dividend_position_ids, transactions, closed_positions, dividends):
    positions = list([x for x in input_positions if x["type"] == typ or (typ == StockType and x['type'] in [FeeType, AdjustmentType, RefundType, IndexAdjustmentType])])
    income_usd = Decimal('0')
    fees_usd = Decimal('0')
    przychod = {}
    koszty = {}
    dochod = {}
    negative_dividend_sum = Decimal('0')
    dividends = group_by_pos_id(dividends)
    refunds_sum_usd = Decimal('0')
    index_adjustment_sum_usd = Decimal('0')

    if unmatched_dividend_position_ids is not None:
        for negative_dividend in [x for x in input_positions if x['type'] == DividendType and x['amount'] < 0 and x['id'] in unmatched_dividend_position_ids]:
            # ujemne dywidendy traktujemy jako koszt, ale tylko dla niezmatchowanych wczesniej dywidend z sheetu 'Dividends'
            pos_id = negative_dividend['id']
            country = get_ticker_country(negative_dividend, transactions, closed_positions, dividends)
            if country == CryptoCountry:
                raise Exception(f"Found a rollover fee for crypto position {pos_id}. Should be marked as cfd?")
            fee = {'id': pos_id, 'date': negative_dividend['date'], 'amount': negative_dividend['amount'], 'country': country, "type": FeeType, 'is_cfd': negative_dividend['is_cfd']}
            positions.append(fee)
            negative_dividend_sum -= negative_dividend['amount']

    for pos in positions:
        pos_id = pos['id']
        country = get_ticker_country(pos, transactions, closed_positions, dividends)
        if country not in przychod:
            przychod[country] = Decimal("0")
            koszty[country] = Decimal("0")
            dochod[country] = Decimal("0")

        if pos["type"] == FeeType:
            rate_pln = convert_rate(pos["date"], pos["amount"], currency='USD')
            fees_usd += pos["amount"]
            if rate_pln > 0:
                # take positive fee for CFD and count it as interest profit
                przychod[country] += rate_pln
            else:
                koszty[country] += -rate_pln
        elif pos["type"] == CryptoType:
            rate_pln = convert_rate(pos["date"], pos["amount"], currency='USD')
            if rate_pln > 0:
                # positive, we sold crypto we bought
                przychod[country] += rate_pln
            else:
                # negative, we bought crypto
                koszty[country] += -rate_pln

            income_usd += pos["equity_change"]
        elif pos['type'] == StockType:
            if pos['is_cfd']:
                profit_pln = convert_rate(pos["close_date"], pos["equity_change"], currency='USD')
                if profit_pln > 0:
                    przychod[country] += profit_pln
                else:
                    koszty[country] += -profit_pln
            else:
                open_rate_pln = convert_rate(pos["open_date"], pos["open_amount"], currency='USD')
                close_rate_pln = convert_rate(pos["close_date"], pos["close_amount"], currency='USD')
                koszty[country] += open_rate_pln
                przychod[country] += close_rate_pln

            income_usd += pos["equity_change"]
        elif pos['type'] in [AdjustmentType, RefundType, IndexAdjustmentType]:
            rate_pln = convert_rate(pos["date"], pos["amount"], currency='USD')
            if rate_pln > 0:
                przychod[country] += rate_pln
            else:
                koszty[country] += -rate_pln

            if pos['type'] in [AdjustmentType, RefundType]:
                refunds_sum_usd += pos['amount']
            elif pos['type'] == IndexAdjustmentType:
                index_adjustment_sum_usd += pos['amount']
            else:
                raise Exception(f'Unknown {pos["type"]} for {pos_id}')
        else:
            raise Exception(f'Unknown {pos["type"]} for {pos_id}')

    for country in dochod.keys():
        dochod[country] = przychod[country] - koszty[country]

    return (income_usd, fees_usd, przychod, koszty, dochod, round(negative_dividend_sum, 2), refunds_sum_usd, index_adjustment_sum_usd)

def process_dividends(incomes, dividend_taxes):
    dividends = [x for x in incomes if x["type"] in [DividendType, InterestType]]
    sum_from_dividend_taxes = sum([item["Net Dividend Received (USD)"] for sublist in dividend_taxes.values() for item in sublist])
    unmatched_dividend_position_ids = set()

    income_dividends_usd = Decimal("0")
    income_dividends_usd2 = Decimal("0")
    income_dividends_usd_brutto = Decimal("0")
    interest_sum_usd = Decimal('0')
    przychod_dywidendy = Decimal("0")
    podatek_nalezny_dywidendy = Decimal("0")
    podatek_zaplacony_dywidendy = Decimal("0")

    for dividend in dividends:
        pos_id = dividend["id"]
        total_usd = dividend["amount"]

        if dividend['type'] == InterestType:
            interest_sum_usd += total_usd
            total_pln = convert_rate(dividend["date"], total_usd, currency='USD')
            przychod_dywidendy += total_pln
            podatek_nalezny_dywidendy += tax_rate * total_pln
            continue
        elif dividend['type'] != DividendType:
            raise Exception("unexpected dividend type")

        if pos_id not in dividend_taxes:
            unmatched_dividend_position_ids.add(pos_id)
            continue

        income_dividends_usd += total_usd
        dividend_tax = next((x for x in dividend_taxes[pos_id] if x["Net Dividend Received (USD)"] == total_usd and x["Date of Payment"].date()== dividend["date"].date()), None)
        if dividend_tax is None:
            raise Exception(f"Unable to match dividend for {pos_id} amount {total_usd} on {dividend['date']}")

        dividend_taxes[pos_id].remove(dividend_tax)
        if len(dividend_taxes[pos_id]) == 0:
            del dividend_taxes[pos_id]

        income_dividends_usd2 += dividend_tax["Net Dividend Received (USD)"]
        witholding_tax_rate = dividend_tax["Withholding Tax Rate (%)"]
        total_usd = dividend_tax["Withholding Tax Amount (USD)"] + dividend_tax["Net Dividend Received (USD)"]
        income_dividends_usd_brutto += total_usd

        total_pln = convert_rate(dividend["date"], total_usd, currency='USD')
        przychod_dywidendy += total_pln
        podatek_zaplacony_dywidendy += witholding_tax_rate * total_pln

        if tax_rate - witholding_tax_rate > 0:
            podatek_nalezny_dywidendy += tax_rate * total_pln
        else:
            podatek_nalezny_dywidendy += witholding_tax_rate * total_pln

    # validate
    if sum_from_dividend_taxes != income_dividends_usd:
        for pos_id, tax in dividend_taxes.items():
            print(f'Pos id: {pos_id} {tax}')
        raise Exception("Suma dywidend między Dywidendy i Transaction Report się nie zgadza. Prawdopodobnie negatywne dywidendy (adjustmenty przez etoro). Zweryfikuj transakcje i manualnie zrób reconciliation w excelu.")

    if len(dividend_taxes) != 0:
        raise Exception("Niewykorzystano wszystkich dywidend do rozdzielenia podatku!")

    income_dividends_usd = round(income_dividends_usd, 2)
    income_dividends_usd_brutto = round(income_dividends_usd_brutto, 4)
    przychod_dywidendy = round(przychod_dywidendy, 4)
    podstawa_dywidendy = round(przychod_dywidendy)
    podatek_nalezny_dywidendy = round(podatek_nalezny_dywidendy)
    podatek_zaplacony_dywidendy = round(podatek_zaplacony_dywidendy)
    return (income_dividends_usd, income_dividends_usd_brutto, przychod_dywidendy, podstawa_dywidendy, podatek_nalezny_dywidendy, podatek_zaplacony_dywidendy, unmatched_dividend_position_ids, interest_sum_usd)

def do_checks(fname, income_dividends_usd, income_stock_usd, fees_stock_usd, negative_dividends, income_crypto_usd, fees_crypto_usd, refunds_sum_usd, interest_sum_usd, index_adjustments_sum_usd):
    stock_sum, crypto_sum, dividends_sum, fees_sum, interest_sum, refunds_sum, index_adjustments_sum = read_summary(fname)
    warnings = []

    if income_dividends_usd != dividends_sum:
        warnings += [f'Dividends check failed. Expected: ${dividends_sum} got {income_dividends_usd}']
    if income_stock_usd != stock_sum:
        warnings += [f'Stock check failed. Expected: ${stock_sum} got {income_stock_usd}']
    if fees_stock_usd + negative_dividends != fees_sum:
        warnings += [f'Fees check failed. Expected: ${fees_sum} got {fees_stock_usd + negative_dividends}']
    if income_crypto_usd != crypto_sum:
        warnings += [f'Crypto check failed. Expected: ${crypto_sum} got {income_crypto_usd}']
    if fees_crypto_usd != Decimal('0'):
        warnings += [f'Crypto check failed. Expected: feed to be 0']
    if refunds_sum != refunds_sum_usd:
        warnings += [f'Incorrect refund sum. Expected ${refunds_sum} got ${refunds_sum_usd}']
    if interest_sum_usd != interest_sum:
        warnings += [f'Incorrect interest sum. Expected ${interest_sum} got ${interest_sum_usd}']
    if index_adjustments_sum_usd != index_adjustments_sum:
        warnings += [f'Incorrect index adjustment sum. Expected ${index_adjustments_sum} got ${index_adjustments_sum_usd}']

    print()
    print("-------------------------IMPORTANT----------------------------")
    if len(warnings) == 0:
        print("Congratulations! All checks with the 'Financial Summary sheet' passed!")
    else:
        [print(x) for x in warnings]
    print("-------------------------IMPORTANT----------------------------")
    print()

fname = 'statement_2024.xlsx'
entries, grouped_transactions, grouped_closed_positions = read(fname)
dividend_taxes, raw_dividends = read_dividend_taxes(fname)

income_dividends_usd, income_dividends_usd_brutto, przychod_dywidendy, podstawa_dywidendy, podatek_nalezny_dywidendy, podatek_zaplacony_dywidendy, unmatched_dividend_position_ids, interest_sum_usd = process_dividends(entries, dividend_taxes)
podatek_do_zaplaty_dywidendy = podatek_nalezny_dywidendy - podatek_zaplacony_dywidendy
print()
print("Dywidendy na PIT-38 sekcja G (podatek poza granicami pln)")
print(f"Przychód $ za dywidendy ${income_dividends_usd} (w summary 'Stock and ETF Dividends (Profit)' + 'CFD Dividends (Profit or Loss)')")
print(f"Przychód $ etoro interest ${interest_sum_usd}")
print(f"Przychód $ za dywidendy brutto ${income_dividends_usd_brutto}")
print(f"Przychód w pln za dywidendy: {przychod_dywidendy} zł")
print(f"Podstawa w pln za dywidendy: {podstawa_dywidendy} zł")
print(f"Podatek należny: {podatek_nalezny_dywidendy} zł")
print(f"Podatek zapłacony za granicą: {podatek_zaplacony_dywidendy} zł")
print(f"Podatek za dywidendy: {podatek_do_zaplaty_dywidendy} zł")

income_stock_usd, fees_stock_usd, przychod_stock, koszty_stock, dochod_stock, negative_dividend_sum, refunds_sum_usd, index_adjustment_sum_usd = process_positions(entries, StockType, unmatched_dividend_position_ids, grouped_transactions, grouped_closed_positions, raw_dividends)
print()
print("Stocks na Pit-38 sekcja C jako inne przychody. Koniecznie z załącznikiem PIT/ZG")
print(f"Dochód $ za stocks: ${income_stock_usd} (w summary suma 'CFDs (Profit or Loss)' + 'Stocks (Profit or Loss)' + 'ETFs (Profit or Loss)')")
print(f"Koszty $ za stocks: ${fees_stock_usd} (w tym negatywne dywidendy: ${negative_dividend_sum}) (w summary suma 'Fees' + 'SDRT Charge' - te negatywne dywidendy czyli ${fees_stock_usd+negative_dividend_sum})")
print(f"Przychód w pln za stocks: {sum_dict(przychod_stock)} zł")
print(f"Koszt w pln za stocks: {sum_dict(koszty_stock)} zł")
print(f"Dochód w pln za stocks: {sum_dict(dochod_stock)} zł")
print(f"Podatek: {max(round(sum_dict(dochod_stock) * tax_rate, 0), 0)} zł")
print(f'Dochód per kraj (zawiera tylko dodatnie): {dict([(x, str(y)) for x, y in dochod_stock.items() if y > 0])}')

income_crypto_usd, fees_crypto_usd, przychod_crypto, koszty_crypto, dochod_crypto, _, _, _ = process_positions(entries, CryptoType, None, grouped_transactions, grouped_closed_positions, raw_dividends)
print()
print("Crypto rozliczamy na PIT-38 sekcja E")
print(f"Dochód $ za crypto: ${income_crypto_usd} (w summary 'Crypto (Profit or Loss)')")
print(f"Koszty $ za crypto: ${fees_crypto_usd} (powinno być zawsze zero)")
print(f"Przychód w pln za crypto: {sum_dict(przychod_crypto)} zł")
print(f"Koszt w pln za crypto: {sum_dict(koszty_crypto)} zł")
print(f"Dochód w pln za crypto: {sum_dict(dochod_crypto)} zł")
print(f"Podatek: {max(round(sum_dict(dochod_crypto) * tax_rate, 0), 0)} zł")

do_checks(fname, income_dividends_usd, income_stock_usd, fees_stock_usd, negative_dividend_sum, income_crypto_usd, fees_crypto_usd, refunds_sum_usd, interest_sum_usd, index_adjustment_sum_usd)